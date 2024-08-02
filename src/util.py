from datetime import datetime
import functools
import uuid

from openpyxl import Workbook, load_workbook
from all_types import SaveUserType, UserType
from constants import DATETIME_FORMAT, DECLARATIONS_SHEET_NAME, SHEETS_TO_PARSE, USERS_SHEET_NAME
from helpers import find_excel_file


def extract_user(usr): return {
    "first_name": usr["firstName"],
    'last_name': usr['lastName'],
    'email': [usr['email']],
    'phone': [usr['phone']]
}


def construct_user_to_save(user: UserType, clerk_user_id: str) -> SaveUserType:
    current_date = datetime.now()

    return {
        **{k: v for k, v in user.items() if k != 'dob'},
        'clerkUserId': clerk_user_id,
        'dob': datetime.strptime(user["dob"], DATETIME_FORMAT),
        'gender': user.get('gender', ''),  # type: ignore
        "createdAt": current_date,
        "updatedAt": current_date,
        "id": uuid.uuid4()
    }  # type: ignore


def extract_sheet(workbook: Workbook, sheet_name: str):
    r"""Extracts a sheet data from an excel sheet

    :param workbook: The workbook object from openpyxl
    :param sheet_name: The sheet from which the data is to be extracted.
    """
    sheet = workbook[sheet_name]
    return list(sheet.values)


@functools.lru_cache()
def get_workbook_and_sheet_names(user_sheet_name: str, declaration_sheet_name: str):
    r"""This function returns a tuple of the workbook, user sheet name,
    declaration sheet name and an array of rest of the fieldnames.

    :param user_sheet_name: The name of the user sheet
    :param declaration_sheet_name: The name of the declaration sheet
    """

    file_to_parse = find_excel_file()

    workbook = load_workbook(filename=file_to_parse, read_only=True)
    fieldnames: list[str] = workbook.sheetnames

    if user_sheet_name != USERS_SHEET_NAME:
        raise ValueError(
            "User details sheet name does not match. Check the constants file.")

    if declaration_sheet_name != DECLARATIONS_SHEET_NAME:
        raise ValueError(
            "Declaration sheet name does not match. Check the constants file.")

    user_sheetname_list = [
        item
        for item in fieldnames
        if item == user_sheet_name
    ]

    if len(user_sheetname_list) == 0:
        raise ValueError("User Details is not present in excel sheet")

    declaration_sheetname_list = [
        item
        for item in fieldnames
        if item == declaration_sheet_name
    ]
    if len(declaration_sheetname_list) == 0:
        raise ValueError(
            "Declarations sheet is not present in excel sheet")

    # exclude User Details and Declaration section
    sheet_names: list[str] = [
        sheet_name for sheet_name in fieldnames
        if sheet_name not in [user_sheet_name, declaration_sheet_name]
        and sheet_name in SHEETS_TO_PARSE
    ]

    return (workbook, user_sheetname_list[0], declaration_sheetname_list[0], sheet_names)
